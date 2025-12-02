from __future__ import annotations

from dataclasses import dataclass

from pathlib import Path
from typing import Dict, List, Tuple, Set

import json
import math
import re

import io
import pandas as pd

from app import config

DATASETS: Dict[str, Dict[str, object]] = {
    "unified": {
        "label": "통합 데이터 (2013~2025)",
        "path": config.SHARED_TOTAL_CSV,
    },
}

# Load term mappings for bilingual search support
_TERM_MAPPINGS: Dict[str, str] = {}
_REVERSE_MAPPINGS: Dict[str, str] = {}

# Load top category mappings for hiding common categories and including them in searches
_TOP_HIDDEN_CATEGORIES: Set[str] = set()
_TOP_CATEGORY_INCLUDES: Dict[str, List[str]] = {}

def _load_top_category_mappings() -> None:
    """Load top category (설비호기) mappings from top_category_mappings.json."""
    global _TOP_HIDDEN_CATEGORIES, _TOP_CATEGORY_INCLUDES

    mappings_path = config.DATA_DIR / "top_category_mappings.json"
    if not mappings_path.exists():
        return

    try:
        with open(mappings_path, 'r', encoding='utf-8') as f:
            data = json.load(f)

        if "hidden_categories" in data:
            _TOP_HIDDEN_CATEGORIES = set(data["hidden_categories"])

        if "category_includes" in data:
            _TOP_CATEGORY_INCLUDES = data["category_includes"]

    except Exception:
        pass

# Load top category mappings at module import
_load_top_category_mappings()

def _load_term_mappings() -> None:
    """Load term mappings from unit_mappings.json for bilingual search."""
    global _TERM_MAPPINGS, _REVERSE_MAPPINGS

    mappings_path = config.DATA_DIR / "unit_mappings.json"
    if not mappings_path.exists():
        return

    try:
        with open(mappings_path, 'r', encoding='utf-8') as f:
            data = json.load(f)

        # Combine all mapping dictionaries
        all_mappings = {}
        if "term_replacements" in data:
            all_mappings.update(data["term_replacements"])
        if "cjk_units" in data:
            all_mappings.update(data["cjk_units"])

        # Normalize keys and values to lowercase
        _TERM_MAPPINGS = {k.lower(): v.lower() for k, v in all_mappings.items()}

        # Create reverse mapping (English -> Korean)
        _REVERSE_MAPPINGS = {v.lower(): k.lower() for k, v in all_mappings.items()}

    except Exception:
        # Silently fail if mapping file is not available
        pass

# Load mappings at module import
_load_term_mappings()

EXCLUDED_TOP_CATEGORIES = {"기계섹션", "예방정비 섹션", "계전섹션", "교육지원섹션"}
MIDDLE_CATEGORY_ALIASES = {
    "기계반": "기계",
    "수산인더스트리 기계": "기계",
    "수산인더스트리-기계": "기계",
    "수산인더스트리_기계": "기계",
    "수산인더스트리기계": "기계",
    "기계반과수산인더스트리 기계": "기계",
    "기계반과수산인더스트리기계": "기계",
    "영진-기계": "기계",
    "영진기계": "기계",
    "영진 기계": "기계",
}
MIDDLE_CATEGORY_ALIASES_NORMALIZED = {
    key.replace(" ", "").replace("-", "").replace("_", ""): value
    for key, value in MIDDLE_CATEGORY_ALIASES.items()
}

COLUMN_ALIASES: Dict[str, str] = {
    "Order Long Text": "정비실적 long text",
    "정비실적 long text": "정비실적 long text",
    "정비실적 Long Text": "정비실적 long text",
    "정비실적 Long text": "정비실적 long text",
    "정비실적  long text": "정비실적 long text",
    "정비실적_long_text": "정비실적 long text",
    "정비실적-Long Text": "정비실적 long text",
    "주요정비실적 Long Text": "정비실적 long text",
    "주요정비실적 long text": "정비실적 long text",
    "정비실적LONGTEXT": "정비실적 long text",
    "정비실적LONG TEXT": "정비실적 long text",
    "정비 실적 long text": "정비실적 long text",
    "정비실적\u3000long text": "정비실적 long text",
    "정비실적\xa0long text": "정비실적 long text",
    "정비실적\u3000Long Text": "정비실적 long text",
    "정비실적\xa0Long Text": "정비실적 long text",
    "정비실적Long Text": "정비실적 long text",
    "정비실적LongText": "정비실적 long text",
    "정비실적_Long Text": "정비실적 long text",
    "정비실적_Long_Text": "정비실적 long text",
    "정비실적-long text": "정비실적 long text",
    "정비실적-long-text": "정비실적 long text",
    "정비실적LONG-TEXT": "정비실적 long text",
    "정비실적_longtext": "정비실적 long text",
    "정비실적long text": "정비실적 long text",
    "정비실적long_text": "정비실적 long text",
    "정비실적 LONG TEXT": "정비실적 long text",
    "정비실적 LONG text": "정비실적 long text",
    "정비실적 LONGText": "정비실적 long text",
    "정비실적 LONGTEXT": "정비실적 long text",
    " long text": "정비실적 long text",
    " long text": "정비실적 long text",
}

MATERIAL_COLUMN_KEY = "materials"

BASE_REQUIRED_COLUMNS = [
    "Order No",
    "Equipment",
    "Order Short Text",
    "Loc. Text",
    "Floc. Text",
    "WorkCtr.Text",
    "Cost Center Text",
    "Object type text",
    "Confirm text",
    "정비실적 long text",
    "Material",
    "Material Desc.",
    "Qty",
    "UoM",
    "Equi. Text",
]

DEFAULT_RESULT_LIMIT = 200

URL_ALLOWED_CHARS = r"A-Za-z0-9\-\._~:/?#\[\]@!$&'()*+,;=%"
URL_PATTERN = re.compile(rf"https?://[{URL_ALLOWED_CHARS}\s]+", re.IGNORECASE)
SEARCH_SELECTION_KEYS = ("top_category", "middle_category", "sub_category", "with_links")
TABLE_COLUMNS = [
    ("dataset_label", "작업일자"),
    ("order_no", "Order No"),
    ("order_short_text", "Order Short Text"),
    ("equipment", "Equipment"),
    ("equi_text", "설비명"),
    ("cost_center", "설비호기"),
    ("workctr", "작업반"),
    ("confirm_text", "정비 Short Text"),
    ("links", "첨부자료"),
    ("details", "상세내역"),
]
SHORT_TEXT_COLUMN = '정비실적 short text'
WORK_DETAIL_FIELDS = [
    ("start_of_execution", "Start of Execution"),
    ("worker_name", "작업자 이름"),
    ("actual_work", "Actual Work"),
    ("work_unit", "Unit"),
]
ORDER_INFO_FIELDS = [
    ("Order No", "Order No"),
    ("Equipment", "Equipment"),
    ("Object type text", "Object type text"),
    ("Order Short Text", "Order Short Text"),
    ("Cost Center Text", "Cost Center Text"),
    ("WorkCtr.Text", "WorkCtr.Text"),
    ("Actual Work", "Actual Work"),
    ("Total Cost", "Total Cost"),
    ("Labor Cost", "Labor Cost"),
    ("Material Cost", "Material Cost"),
    ("Other Cost", "Other Cost"),
]



@dataclass
class DataStore:
    combined: pd.DataFrame
    top_options: List[str]
    middle_options: Dict[str, List[str]]
    sub_options: Dict[Tuple[str, str], List[str]]
    all_middle_options: List[str]
    sub_by_middle: Dict[str, List[str]]
    sub_by_top: Dict[str, List[str]]
    all_sub_options: List[str]


DATA_STORE: DataStore | None = None
DATASET_MTIMES: Dict[str, float] = {}


def _resolve_limit(raw_value: str | None) -> int:
    try:
        value = int(raw_value) if raw_value else DEFAULT_RESULT_LIMIT
    except (TypeError, ValueError):
        value = DEFAULT_RESULT_LIMIT
    if value < DEFAULT_RESULT_LIMIT:
        value = DEFAULT_RESULT_LIMIT
    return value


def _capture_dataset_mtimes() -> Dict[str, float]:
    mtimes: Dict[str, float] = {}
    for key, config in DATASETS.items():
        path = config["path"]
        try:
            mtimes[key] = path.stat().st_mtime
        except FileNotFoundError:
            mtimes[key] = 0.0
    return mtimes


def _clean_string_series(series: pd.Series) -> pd.Series:
    return (
        series.replace({pd.NA: "", "nan": "", "NaN": "", None: ""})
        .astype(str)
        .str.strip()
    )


def _apply_column_aliases(df: pd.DataFrame) -> pd.DataFrame:
    result = df.copy()

    for source, target in COLUMN_ALIASES.items():
        if source not in result.columns:
            continue
        if source == target:
            continue

        if target in result.columns:
            source_clean = _clean_string_series(result[source])
            target_clean = _clean_string_series(result[target])
            mask = target_clean == ""
            if mask.any():
                result.loc[mask, target] = source_clean[mask]
            result.drop(columns=[source], inplace=True)
        else:
            result.rename(columns={source: target}, inplace=True)

    return result


def _read_dataset(path: Path) -> pd.DataFrame:
    print(f"[data_store] _read_dataset called with: {path}")
    print(f"[data_store] File exists: {path.exists()}, suffix: {path.suffix}")

    if not path.exists():
        print(f"[data_store] ERROR: File does not exist!")
        return pd.DataFrame(columns=BASE_REQUIRED_COLUMNS)

    # Check if file is SQLite database
    if path.suffix.lower() == '.db':
        print(f"[data_store] Reading SQLite database...")
        import sqlite3
        try:
            conn = sqlite3.connect(str(path))
            df = pd.read_sql_query("SELECT * FROM sap_reports", conn, dtype=str)
            conn.close()
            print(f"[data_store] Loaded {len(df)} rows from database")
        except Exception as exc:
            print(f"[data_store] ERROR reading database: {exc}")
            raise RuntimeError(f"Failed to read SQLite database {path}: {exc}")
    else:
        # Read CSV file with encoding fallback
        encodings = ("utf-8-sig", "cp949", "utf-8")
        last_error: Exception | None = None
        for encoding in encodings:
            try:
                df = pd.read_csv(path, dtype=str, low_memory=False, encoding=encoding)
                break
            except UnicodeDecodeError as exc:
                last_error = exc
        else:
            if last_error is not None:
                raise last_error
            df = pd.DataFrame(columns=BASE_REQUIRED_COLUMNS)

    df.columns = [str(column).strip() for column in df.columns]
    if df.columns.duplicated().any():
        df = df.loc[:, ~df.columns.duplicated()]
    if COLUMN_ALIASES:
        df = _apply_column_aliases(df)
        if df.columns.duplicated().any():
            df = df.loc[:, ~df.columns.duplicated()]

    df = df.replace({pd.NA: "", "nan": "", "NaN": ""}).fillna("")

    for column in BASE_REQUIRED_COLUMNS:
        if column not in df.columns:
            df[column] = ""
        df[column] = df[column].astype(str).str.strip()

    # Remove decimal points from numeric fields (e.g., 1008483.0 -> 1008483)
    numeric_fields = ["Order No", "Equipment", "Man", "Actual Duration", "Actual Work", "Material", "Qty", "Noti. No"]
    for field in numeric_fields:
        if field in df.columns:
            df[field] = df[field].astype(str).str.replace(r"\.0$", "", regex=True).str.strip()

    return df


def _alias_middle_value(value: str) -> str:
    key = value.strip()
    if not key:
        return ""
    normalized = key.replace(" ", "").replace("-", "").replace("_", "")
    return MIDDLE_CATEGORY_ALIASES.get(
        key,
        MIDDLE_CATEGORY_ALIASES_NORMALIZED.get(normalized, key),
    )


def _calculate_work_date_for_sort(df: pd.DataFrame) -> pd.Series:
    """Calculate WorkDateForSort once for all data. Returns a Series indexed by Order No."""
    if df.empty or "Order No" not in df.columns:
        return pd.Series(dtype=str)

    # Try date columns in order of preference
    date_columns = ["Start of Execution", "Bsc start", "Actual Start (Time)", "Required Start"]
    order_dates = pd.Series(index=df["Order No"].unique(), dtype=str)
    order_dates[:] = ""

    for col in date_columns:
        if col not in df.columns:
            continue

        # Clean date column
        date_series = (
            df[col]
            .fillna("")
            .astype(str)
            .str.strip()
            .replace({"None": "", "nan": "", "NaN": "", "nat": "", "NaT": ""})
            .str.split(" ").str[0]  # Remove time portion
        )

        # Get rows with valid dates
        valid_mask = date_series != ""
        if not valid_mask.any():
            continue

        # Get minimum date per Order No for orders without dates yet
        temp_df = df[valid_mask][["Order No"]].copy()
        temp_df["_date"] = date_series[valid_mask]
        grouped = temp_df.groupby("Order No")["_date"].min()

        # Update only orders that don't have dates yet
        needs_date = order_dates == ""
        order_dates[needs_date] = grouped.reindex(order_dates[needs_date].index).fillna("")

        # If all orders have dates, stop
        if (order_dates != "").all():
            break

    return order_dates


def _add_alias_columns(df: pd.DataFrame) -> pd.DataFrame:
    # Work directly on dataframe without copy
    df["WorkCtrAlias"] = df["WorkCtr.Text"].apply(_alias_middle_value)
    df["OrderNoNumeric"] = pd.to_numeric(df["Order No"], errors="coerce")
    long_text_series = df["정비실적 long text"].astype(str).fillna("")
    df["HasLongTextLink"] = long_text_series.apply(
        lambda value: bool(URL_PATTERN.search(value))
    )

    # WorkDateForSort will be added once during data loading
    # This function no longer calculates it
    if "WorkDateForSort" not in df.columns:
        df["WorkDateForSort"] = ""

    return df


def _select_order_numbers(
    filtered: pd.DataFrame, limit: int | None = None
) -> List[str]:
    if filtered.empty:
        return []

    if limit is None:
        limit = DEFAULT_RESULT_LIMIT

    # Get one row per order with the work date and Order Short Text
    order_keys = (
        filtered[["Order No", "OrderNoNumeric", "WorkDateForSort", "Order Short Text"]]
        .drop_duplicates(subset="Order No", keep="first")
        .assign(OrderNoNumeric=lambda df: df["OrderNoNumeric"].fillna(-math.inf))
    )

    # Add priority flag: orders with "도면정보" in Order Short Text get highest priority
    order_keys["HasDrawingInfo"] = order_keys["Order Short Text"].astype(str).str.contains("도면정보", case=False, na=False)

    # Separate into three groups: with drawing info, with dates, without dates
    has_drawing_info = order_keys["HasDrawingInfo"]
    drawing_orders = order_keys[has_drawing_info].copy()
    non_drawing_orders = order_keys[~has_drawing_info].copy()

    # Sort drawing info orders by date (descending), then by order number
    if not drawing_orders.empty:
        has_date_drawing = drawing_orders["WorkDateForSort"] != ""
        with_dates_drawing = drawing_orders[has_date_drawing].copy()
        without_dates_drawing = drawing_orders[~has_date_drawing].copy()

        if not with_dates_drawing.empty:
            with_dates_drawing = with_dates_drawing.sort_values(
                by=["WorkDateForSort", "OrderNoNumeric", "Order No"],
                ascending=[False, False, False]
            )
        if not without_dates_drawing.empty:
            without_dates_drawing = without_dates_drawing.sort_values(
                by=["OrderNoNumeric", "Order No"],
                ascending=[False, False]
            )

        if not with_dates_drawing.empty and not without_dates_drawing.empty:
            drawing_orders = pd.concat([with_dates_drawing, without_dates_drawing], ignore_index=True)
        elif not with_dates_drawing.empty:
            drawing_orders = with_dates_drawing
        else:
            drawing_orders = without_dates_drawing

    # Sort non-drawing orders by date (descending), then by order number
    if not non_drawing_orders.empty:
        has_date = non_drawing_orders["WorkDateForSort"] != ""
        with_dates = non_drawing_orders[has_date].copy()
        without_dates = non_drawing_orders[~has_date].copy()

        if not with_dates.empty:
            with_dates = with_dates.sort_values(
                by=["WorkDateForSort", "OrderNoNumeric", "Order No"],
                ascending=[False, False, False]
            )
        if not without_dates.empty:
            without_dates = without_dates.sort_values(
                by=["OrderNoNumeric", "Order No"],
                ascending=[False, False]
            )

        if not with_dates.empty and not without_dates.empty:
            non_drawing_orders = pd.concat([with_dates, without_dates], ignore_index=True)
        elif not with_dates.empty:
            non_drawing_orders = with_dates
        else:
            non_drawing_orders = without_dates

    # Combine: drawing info orders first, then other orders
    if not drawing_orders.empty and not non_drawing_orders.empty:
        combined = pd.concat([drawing_orders, non_drawing_orders], ignore_index=True)
    elif not drawing_orders.empty:
        combined = drawing_orders
    else:
        combined = non_drawing_orders

    return combined["Order No"].head(limit).tolist()


def _initialize_data() -> DataStore:
    frames: List[pd.DataFrame] = []

    for dataset_key, config in DATASETS.items():
        df = _read_dataset(config["path"])
        if df.empty:
            continue
        df["dataset_key"] = dataset_key
        df["dataset_label"] = config["label"]
        df = _add_alias_columns(df)
        frames.append(df)

    if frames:
        combined = pd.concat(frames, ignore_index=True)
    else:
        combined = pd.DataFrame(columns=BASE_REQUIRED_COLUMNS + ["dataset_key", "dataset_label"])
        combined = _add_alias_columns(combined)

    # Calculate WorkDateForSort once for all data
    if not combined.empty:
        print("[data_store] Calculating WorkDateForSort...")
        order_dates = _calculate_work_date_for_sort(combined)
        combined["WorkDateForSort"] = combined["Order No"].map(order_dates).fillna("")
        print(f"[data_store] WorkDateForSort calculated for {len(order_dates)} orders")

    # Filter orders: must have at least one of these fields populated
    # 1. 정비실적 short text (정비 Short Text)
    # 2. 정비실적 long text (Long Text)
    # 3. Material (자재 정보)

    if SHORT_TEXT_COLUMN in combined.columns:
        short_text = combined[SHORT_TEXT_COLUMN].astype(str).str.strip()
        # Filter out 'nan', 'None' and empty strings
        short_text = short_text.replace({"nan": "", "NaN": "", "None": ""})
    else:
        short_text = pd.Series([""] * len(combined), index=combined.index)

    if "정비실적 long text" in combined.columns:
        long_text = combined["정비실적 long text"].astype(str).str.strip()
        # Filter out 'nan', 'None' and empty strings
        long_text = long_text.replace({"nan": "", "NaN": "", "None": ""})
    else:
        long_text = pd.Series([""] * len(combined), index=combined.index)

    if "Material" in combined.columns:
        material = combined["Material"].astype(str).str.strip()
        # Filter out 'nan' and empty strings
        material = material.replace({"nan": "", "NaN": "", "None": ""})
    else:
        material = pd.Series([""] * len(combined), index=combined.index)

    # Row has required data if ANY of these three fields is non-empty
    has_required_data = (short_text != "") | (long_text != "") | (material != "")

    if "Order No" in combined.columns:
        # Order is valid if ANY row in that order has required data
        order_has_data = has_required_data.groupby(combined["Order No"]).transform("any")
        combined = combined[order_has_data].copy()
    else:
        combined = combined[has_required_data].copy()

    top_candidates = (
        combined["Cost Center Text"].astype(str).str.strip().replace({"nan": "", "NaN": ""})
    )
    # Exclude both EXCLUDED_TOP_CATEGORIES and hidden categories from mappings
    excluded_set = EXCLUDED_TOP_CATEGORIES | _TOP_HIDDEN_CATEGORIES
    all_top_values = sorted(
        {
            value
            for value in top_candidates
            if value and value not in excluded_set
        }
    )

    # Priority ordering: 인천복합발전3호기~9호기 first, then separator, then others
    priority_items = [
        "인천복합발전3호기",
        "인천복합발전4호기",
        "인천복합발전5호기",
        "인천복합발전6호기",
        "인천복합발전7호기",
        "인천복합발전8호기",
        "인천복합발전9호기",
    ]

    # Build ordered list: priority items first (if they exist), then separators, then rest
    top_options = []
    for item in priority_items:
        if item in all_top_values:
            top_options.append(item)

    # Add 15 empty separator items (for visual spacing in dropdown)
    for _ in range(15):
        top_options.append("─" * 20)  # Visual separator line

    # Add remaining items (excluding priority items)
    for item in all_top_values:
        if item not in priority_items:
            top_options.append(item)

    middle_options: Dict[str, List[str]] = {}
    all_middle_set: set[str] = set()
    sub_by_middle: Dict[str, set[str]] = {}
    sub_by_top: Dict[str, set[str]] = {}

    for top in top_options:
        if top.startswith("─"):  # Skip separator lines
            continue
        aliases = (
            combined.loc[combined["Cost Center Text"] == top, "WorkCtrAlias"]
            .dropna()
            .astype(str)
            .str.strip()
        )
        alias_list = sorted({alias for alias in aliases if alias})
        middle_options[top] = alias_list
        all_middle_set.update(alias_list)

    sub_options: Dict[Tuple[str, str], List[str]] = {}
    for top in top_options:
        if top.startswith("─"):  # Skip separator lines
            continue
        top_subs: set[str] = set()
        for alias in middle_options.get(top, []):
            sub_vals = (
                combined.loc[
                    (combined["Cost Center Text"] == top)
                    & (combined["WorkCtrAlias"] == alias),
                    "Object type text",
                ]
                .dropna()
                .astype(str)
                .str.strip()
            )
            clean_subs = {value for value in sub_vals if value}
            sub_options[(top, alias)] = sorted(clean_subs)
            if clean_subs:
                sub_by_middle.setdefault(alias, set()).update(clean_subs)
                top_subs.update(clean_subs)
        sub_by_top[top] = top_subs

    all_middle_options = sorted(all_middle_set)
    all_sub_options = sorted(
        {
            value
            for value in combined["Object type text"].astype(str).str.strip()
            if value
        }
    )
    sub_by_middle_lists = {alias: sorted(values) for alias, values in sub_by_middle.items()}
    sub_by_top_lists = {top: sorted(values) for top, values in sub_by_top.items()}

    return DataStore(
        combined=combined,
        top_options=top_options,
        middle_options=middle_options,
        sub_options=sub_options,
        all_middle_options=all_middle_options,
        sub_by_middle=sub_by_middle_lists,
        sub_by_top=sub_by_top_lists,
        all_sub_options=all_sub_options,
    )


def _get_data_store() -> DataStore:
    global DATA_STORE
    current_mtimes = _capture_dataset_mtimes()
    needs_reload = DATA_STORE is None or any(
        DATASET_MTIMES.get(key) != mtime for key, mtime in current_mtimes.items()
    )

    if needs_reload:
        DATA_STORE = _initialize_data()
        DATASET_MTIMES.clear()
        DATASET_MTIMES.update(current_mtimes)
    return DATA_STORE


def _extract_word_tokens(text: str) -> List[str]:
    """Extract individual words from text as normalized tokens.

    Splits text by special characters and whitespace, normalizes each word.
    Each token represents an independent word.

    Example:
        "SLP-C Pump" -> ["slp", "c", "pump"]
        "SLP Screen Wash Pump C COUPLING" -> ["slp", "screen", "wash", "pump", "c", "coupling"]
        "[정산] CC#56 SLP-C 외주" -> ["정산", "cc", "56", "slp", "c", "외주"]

    Args:
        text: The text to extract tokens from

    Returns:
        List of normalized word tokens
    """
    if pd.isna(text):
        return []
    # Extract all continuous alphanumeric/Korean sequences (splits by special chars and spaces)
    tokens = re.findall(r'[a-z0-9가-힣]+', str(text).lower())
    return tokens


def _expand_search_tokens_with_mappings(search_tokens: List[str]) -> List[Set[str]]:
    """Expand search tokens with their bilingual equivalents from mappings.

    For each search token, creates a set containing:
    - The original token
    - Its mapped equivalent (if exists in _TERM_MAPPINGS or _REVERSE_MAPPINGS)

    Example:
        ["pump"] -> [{"pump", "펌프"}]
        ["펌프"] -> [{"펌프", "pump"}]
        ["slp", "c"] -> [{"slp"}, {"c"}]

    Args:
        search_tokens: List of search tokens

    Returns:
        List of sets, each containing a token and its equivalents
    """
    expanded = []
    for token in search_tokens:
        token_set = {token}

        # Check Korean -> English mapping
        if token in _TERM_MAPPINGS:
            token_set.add(_TERM_MAPPINGS[token])

        # Check English -> Korean mapping
        if token in _REVERSE_MAPPINGS:
            token_set.add(_REVERSE_MAPPINGS[token])

        expanded.append(token_set)

    return expanded


def _contains_all_word_tokens(text_tokens: List[str], search_token_sets: List[Set[str]]) -> bool:
    """Check if all search token sets have at least one match in text tokens.

    This supports bilingual search where each search token can match multiple equivalents.

    This ensures "slp c" matches:
    - "SLP-C" -> ["slp", "c"] ✓ (both exist as independent words)
    - "SLP Screen C" -> ["slp", "screen", "c"] ✓ (both exist as independent words)

    And "pump" matches:
    - "Feed Water Pump" -> [..., "pump"] ✓
    - "급수 펌프" -> [..., "펌프"] ✓ (matched via mapping)

    But NOT:
    - "SLP COUPLING" -> ["slp", "coupling"] ✗ ("c" doesn't exist as independent word)

    Args:
        text_tokens: List of word tokens from the text
        search_token_sets: List of sets, each containing equivalent tokens

    Returns:
        True if all search token sets have at least one match
    """
    for token_set in search_token_sets:
        # Check if any token in the set exists in text_tokens
        if not any(token in text_tokens for token in token_set):
            return False
    return True


def _apply_filters(df: pd.DataFrame, selections: Dict[str, str]) -> pd.DataFrame:
    filtered = df

    equipment_no = selections.get("equipment_no", "").strip()
    if equipment_no:
        filtered = filtered[
            filtered["Equipment"].str.contains(equipment_no, case=False, na=False)
        ]

    order_no = selections.get("order_no", "").strip()
    if order_no:
        filtered = filtered[
            filtered["Order No"].str.contains(order_no, case=False, na=False)
        ]

    equipment_name = selections.get("equipment_name", "").strip()
    if equipment_name:
        # Extract search tokens (independent words)
        search_tokens = _extract_word_tokens(equipment_name)

        if search_tokens:
            # Expand search tokens with bilingual mappings
            search_token_sets = _expand_search_tokens_with_mappings(search_tokens)

            # Extract word tokens from each field
            order_short_tokens = filtered["Order Short Text"].apply(_extract_word_tokens)
            equi_text_tokens = filtered["Equi. Text"].apply(_extract_word_tokens)

            # Check if all search token sets have matches in either field
            order_short_match = order_short_tokens.apply(
                lambda tokens: _contains_all_word_tokens(tokens, search_token_sets)
            )
            equi_text_match = equi_text_tokens.apply(
                lambda tokens: _contains_all_word_tokens(tokens, search_token_sets)
            )
            filtered = filtered[order_short_match | equi_text_match]

    top_category = selections.get("top_category", "").strip()
    # Skip separator lines (they shouldn't be selectable, but just in case)
    if top_category and not top_category.startswith("─"):
        # Include common categories that should be included with this top category
        categories_to_include = [top_category]
        if top_category in _TOP_CATEGORY_INCLUDES:
            categories_to_include.extend(_TOP_CATEGORY_INCLUDES[top_category])
        filtered = filtered[filtered["Cost Center Text"].isin(categories_to_include)]

    middle_category = selections.get("middle_category", "").strip()
    if middle_category:
        filtered = filtered[filtered["WorkCtrAlias"] == middle_category]

    sub_category = selections.get("sub_category", "").strip()
    if sub_category:
        filtered = filtered[filtered["Object type text"] == sub_category]

    with_links = selections.get("with_links", "").strip() == "1"
    if with_links and "HasLongTextLink" in filtered.columns:
        filtered = filtered[filtered["HasLongTextLink"]]

    # Detail query filter: search in Material, Material Desc., and 작업자 이름
    detail_query = selections.get("detail_query", "").strip()
    if detail_query:
        # Create a mask for rows that match the detail query
        material_match = filtered["Material"].astype(str).str.contains(detail_query, case=False, na=False)
        material_desc_match = filtered["Material Desc."].astype(str).str.contains(detail_query, case=False, na=False)
        worker_match = filtered["작업자 이름"].astype(str).str.contains(detail_query, case=False, na=False) if "작업자 이름" in filtered.columns else pd.Series([False] * len(filtered), index=filtered.index)
        
        # Combine masks: row matches if ANY of the columns contain the query
        row_matches = material_match | material_desc_match | worker_match
        
        # Get Order Nos that have at least one matching row
        matching_orders = filtered.loc[row_matches, "Order No"].unique()
        
        # Filter to only include orders that have at least one matching row
        filtered = filtered[filtered["Order No"].isin(matching_orders)]

    return filtered


def _unique_preserve(values: pd.Series) -> List[str]:
    seen: set[str] = set()
    ordered: List[str] = []
    for raw in values.astype(str):
        value = raw.strip()
        # Filter out "None", "nan" strings
        if not value or value.lower() in ("none", "nan", "nat") or value in seen:
            continue
        seen.add(value)
        ordered.append(value)
    return ordered


def _extract_links(text: str) -> Tuple[str, List[str]]:
    if not text:
        return "", []

    links: List[str] = []

    def _replace(match: re.Match[str]) -> str:
        raw = match.group(0)
        cleaned = re.sub(r"\s+", "", raw)
        if cleaned:
            links.append(cleaned)

        preserved_newlines = "".join(char for char in raw if char in ("\n", "\r"))
        if preserved_newlines:
            return preserved_newlines
        return " "

    cleaned_text = URL_PATTERN.sub(_replace, text)
    cleaned_text = cleaned_text.strip()
    return cleaned_text, links


def _format_work_date_label(row: pd.Series) -> str:
    """Get work date label for a single row.

    Uses the WorkDateForSort column which contains the minimum (earliest) date
    for the order, ensuring consistency between sorting and display.
    """
    # Use WorkDateForSort which already contains the earliest date for the order
    work_date = row.get("WorkDateForSort", "")
    if work_date and str(work_date).strip() and str(work_date).strip().lower() not in ("none", "nan", "nat", ""):
        return str(work_date).strip()

    # Return empty string if no date found
    return ""


def _format_work_date_label_for_group(group: pd.DataFrame) -> str:
    """Get work date label for an entire order group.

    Uses the WorkDateForSort column which contains the minimum (earliest) date
    for the order, ensuring all rows in the same order show the same work date
    and it matches the sorting criteria.
    """
    if group.empty:
        return ""

    # Use WorkDateForSort from first row (all rows in same order have same value)
    first = group.iloc[0]
    work_date = first.get("WorkDateForSort", "")
    if work_date and str(work_date).strip() and str(work_date).strip().lower() not in ("none", "nan", "nat", ""):
        return str(work_date).strip()

    # Return empty string if no date found
    return ""


def _collect_confirm_texts(group: pd.DataFrame) -> List[str]:
    columns = ["Confirm text"]
    if SHORT_TEXT_COLUMN in group.columns:
        columns.append(SHORT_TEXT_COLUMN)

    seen: set[str] = set()
    results: List[str] = []

    for column in columns:
        if column not in group.columns:
            continue
        for value in _unique_preserve(group[column]):
            value = value.strip()
            if not value or value.lower() in {"nan", "none"} or value in seen:
                continue
            seen.add(value)
            results.append(value)
    return results


def _collect_work_details(group: pd.DataFrame) -> List[Dict[str, str]]:
    entries: List[Dict[str, str]] = []
    seen: set[tuple[str, ...]] = set()

    for _, row in group.iterrows():
        entry: Dict[str, str] = {}
        values: List[str] = []
        has_value = False
        for key, column in WORK_DETAIL_FIELDS:
            raw = row.get(column, "")
            value = "" if pd.isna(raw) else str(raw).strip()
            # Filter out "None", "nan" strings
            if value.lower() in ("none", "nan", "nat"):
                value = ""
            # Remove time portion from Start of Execution (00:00:00)
            if key == "start_of_execution" and value:
                value = value.split(" ")[0]
            if key == "actual_work" and value:
                normalized = value.replace(",", "")
                try:
                    if float(normalized or "0") == 0:
                        value = ""
                except ValueError:
                    pass
            entry[key] = value
            values.append(value)
            if value:
                has_value = True
        if not has_value:
            continue
        entry_key = tuple(values)
        if entry_key in seen:
            continue
        seen.add(entry_key)
        entries.append(entry)
    return entries


def _build_table_rows(filtered: pd.DataFrame, selected_orders: List[str] | None = None) -> List[Dict[str, object]]:
    if filtered.empty:
        return []

    if selected_orders is None:
        selected_orders = _select_order_numbers(filtered)

    grouped = filtered.groupby("Order No", sort=False)
    rows: List[Dict[str, object]] = []

    for order_no in selected_orders:
        if order_no not in grouped.groups:
            continue

        group = grouped.get_group(order_no).copy()
        group.sort_index(inplace=True)
        first = group.iloc[0]

        confirm_values = _collect_confirm_texts(group)
        confirm_value = "\n".join(confirm_values)

        long_text_values = _unique_preserve(group["정비실적 long text"])
        long_text_parts: List[str] = []
        long_links: List[str] = []
        for raw_value in long_text_values:
            cleaned, links = _extract_links(raw_value)
            if cleaned:
                long_text_parts.append(cleaned)
            long_links.extend(links)
        long_links = list(dict.fromkeys(long_links))
        long_text_combined = "\n".join(long_text_parts).strip()

        materials_df = (
            group[["Material", "Material Desc.", "Qty", "UoM"]]
            .drop_duplicates()
            .reset_index(drop=True)
        )
        material_entries: List[Dict[str, str]] = []
        for _, material_row in materials_df.iterrows():
            entry = {
                "material": str(material_row["Material"]).strip(),
                "description": str(material_row["Material Desc."]).strip(),
                "qty": str(material_row["Qty"]).strip(),
                "uom": str(material_row["UoM"]).strip(),
            }
            # Filter out "None", "nan" strings
            for key in entry:
                if entry[key].lower() in ("none", "nan", "nat"):
                    entry[key] = ""

            has_text = bool(entry["material"] or entry["description"])
            qty_value = entry["qty"].replace(",", "")
            has_qty = False
            if qty_value:
                try:
                    has_qty = float(qty_value) > 0
                except ValueError:
                    has_qty = True
            if has_text or has_qty:
                material_entries.append(entry)

        work_date_label = _format_work_date_label_for_group(group)
        work_details = _collect_work_details(group)

        detail_payload = {
            "order_no": order_no,
            "dataset_label": work_date_label,
            "long_text": long_text_combined,
            "long_text_links": long_links,
            "materials": material_entries,
            "work_details": work_details,
        }

        rows.append(
            {
                "dataset_label": work_date_label,
                "order_no": order_no,
                "order_short_text": first.get("Order Short Text", ""),
                "equipment": first.get("Equipment", ""),
                "equi_text": first.get("Equi. Text", ""),
                "workctr": first.get("WorkCtr.Text", ""),
                "cost_center": first.get("Cost Center Text", ""),
                "confirm_text": confirm_value,
                "has_links": bool(long_links),
                "links": long_links,
                "long_text": long_text_combined,
                MATERIAL_COLUMN_KEY: material_entries,
                "detail_payload": detail_payload,
            }
        )

    return rows


def build_excel_export_data(table_rows: List[Dict[str, object]]) -> pd.DataFrame:
    """Convert table rows to flattened DataFrame for Excel export.

    Includes basic info + work details + long text + materials from detail_payload.
    """
    if not table_rows:
        return pd.DataFrame()

    export_rows: List[Dict[str, str]] = []

    for row in table_rows:
        detail_payload = row.get("detail_payload", {})
        work_details = detail_payload.get("work_details", [])
        materials = detail_payload.get("materials", [])
        long_text = detail_payload.get("long_text", "")

        # Base row data
        base_data = {
            "작업일자": str(row.get("dataset_label", "")),
            "Order No": str(row.get("order_no", "")),
            "Order Short Text": str(row.get("order_short_text", "")),
            "Equipment": str(row.get("equipment", "")),
            "설비명": str(row.get("equi_text", "")),
            "설비호기": str(row.get("cost_center", "")),
            "작업반": str(row.get("workctr", "")),
            "정비 Short Text": str(row.get("confirm_text", "")),
        }

        # Determine max count for work details and materials
        max_work_count = len(work_details)
        max_material_count = len(materials)

        # If there's no work details or materials, create one row with just base data
        if max_work_count == 0 and max_material_count == 0:
            export_row = base_data.copy()
            export_row["정비실적 Long Text"] = long_text
            export_rows.append(export_row)
            continue

        # Create rows for each combination
        max_count = max(max_work_count, max_material_count, 1)

        for i in range(max_count):
            export_row = base_data.copy()

            # Add work detail columns
            if i < max_work_count:
                work = work_details[i]
                export_row["작업 시작일"] = work.get("start_of_execution", "")
                export_row["작업자 이름"] = work.get("worker_name", "")
                export_row["작업 시간"] = work.get("actual_work", "")
                export_row["작업 시간 단위"] = work.get("work_unit", "")
            else:
                export_row["작업 시작일"] = ""
                export_row["작업자 이름"] = ""
                export_row["작업 시간"] = ""
                export_row["작업 시간 단위"] = ""

            # Add long text (only on first row to avoid repetition)
            if i == 0:
                export_row["정비실적 Long Text"] = long_text
            else:
                export_row["정비실적 Long Text"] = ""

            # Add material columns
            if i < max_material_count:
                material = materials[i]
                export_row["자재 코드"] = material.get("material", "")
                export_row["자재 설명"] = material.get("description", "")
                export_row["수량"] = material.get("qty", "")
                export_row["단위"] = material.get("uom", "")
            else:
                export_row["자재 코드"] = ""
                export_row["자재 설명"] = ""
                export_row["수량"] = ""
                export_row["단위"] = ""

            export_rows.append(export_row)

    return pd.DataFrame(export_rows)


def insert_blank_rows_between_orders(df: pd.DataFrame) -> pd.DataFrame:
    """Insert blank rows between different Order Nos in the DataFrame.

    Args:
        df: DataFrame with 'Order No' column

    Returns:
        DataFrame with blank rows inserted between different order numbers
    """
    if df.empty or "Order No" not in df.columns:
        return df

    result_rows: List[pd.Series] = []
    prev_order_no = None

    for idx, row in df.iterrows():
        current_order_no = row["Order No"]

        # Insert blank row if order number changed (but not for the first row)
        if prev_order_no is not None and current_order_no != prev_order_no:
            # Create blank row with same columns
            blank_row = pd.Series([""] * len(df.columns), index=df.columns)
            result_rows.append(blank_row)

        result_rows.append(row)
        prev_order_no = current_order_no

    return pd.DataFrame(result_rows).reset_index(drop=True)


def format_excel_worksheet(worksheet, df: pd.DataFrame) -> None:
    """Format Excel worksheet for better readability.

    Adjusts column widths and row heights to make content more visible.

    Args:
        worksheet: openpyxl worksheet object
        df: DataFrame that was written to the worksheet
    """
    from openpyxl.styles import Alignment

    # Auto-adjust column widths based on content
    for idx, column in enumerate(df.columns, start=1):
        column_letter = worksheet.cell(row=1, column=idx).column_letter

        # Calculate max width for this column
        max_length = len(str(column))  # Start with header length

        for row_idx in range(2, len(df) + 2):  # +2 because Excel is 1-indexed and has header
            cell_value = worksheet.cell(row=row_idx, column=idx).value
            if cell_value:
                # Count newlines for multi-line content
                cell_str = str(cell_value)
                lines = cell_str.split('\n')
                # Get the longest line
                line_length = max(len(line) for line in lines) if lines else 0
                max_length = max(max_length, line_length)

        # Set column width with a reasonable limit (between 10 and 80)
        adjusted_width = min(max(max_length + 2, 10), 80)
        worksheet.column_dimensions[column_letter].width = adjusted_width

    # Set row heights and enable text wrapping
    for row_idx in range(1, len(df) + 2):  # +2 for header and 1-indexing
        # Set row height (default Excel row height is ~15, we use 25 for better readability)
        worksheet.row_dimensions[row_idx].height = 25

        # Enable text wrapping for all cells in this row
        for col_idx in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
